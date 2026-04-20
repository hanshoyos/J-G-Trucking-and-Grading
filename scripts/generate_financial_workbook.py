from __future__ import annotations

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.workbook.defined_name import DefinedName

OUTPUT = "Personal Financial Master Plan.xlsx"

SHEETS = [
    "Master Data",
    "Dashboard",
    "Paycheck Calculator",
    "Tax Planning",
    "Income",
    "Budget",
    "Debt Tracker",
    "Debt Payoff Strategy",
    "Emergency Fund Analysis",
    "Child Support Projection",
    "Debt vs Investment Analysis",
    "200K Allocation Plan",
    "Tax-Optimized Savings Strategy",
    "Cash Flow Waterfall",
    "Scenario Modeling",
    "Retirement Planner",
    "Investment Allocation Tracker",
    "Child Support Impact Analysis",
    "Net Worth Statement",
    "Monthly Update Checklist",
    "Settings",
    "Tax Assumptions",
]

COLORS = {
    "blue_text": "000000FF",
    "black_text": "00000000",
    "green_text": "00008000",
    "yellow_fill": "00FFFF00",
    "red_text": "00FF0000",
    "gray_fill": "00F2F2F2",
    "title_fill": "001F4E78",
    "header_fill": "004F81BD",
}

TAB_COLORS = {
    "blue": "4F81BD",
    "red": "C0504D",
    "green": "9BBB59",
    "purple": "8064A2",
    "gold": "D8B200",
    "gray": "7F7F7F",
}


def title(ws, text):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=16, color="00FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["title_fill"])
    ws.row_dimensions[1].height = 26


def section(ws, r, text):
    ws[f"A{r}"] = text
    ws[f"A{r}"].font = Font(bold=True, size=12, color="00FFFFFF")
    ws[f"A{r}"].fill = PatternFill("solid", fgColor=COLORS["header_fill"])
    ws.row_dimensions[r].height = 24


def set_layout(ws):
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 34
    for c in "BCDEFGHIJKLMNOPQRSTUVWXYZ":
        ws.column_dimensions[c].width = 16
    ws.sheet_view.showGridLines = False


def input_cell(cell):
    cell.font = Font(color=COLORS["blue_text"])


def formula_cell(cell, cross=False):
    cell.font = Font(color=COLORS["green_text"] if cross else COLORS["black_text"])


def currency(cell):
    cell.number_format = '$#,##0_);($#,##0);-'


def percent(cell):
    cell.number_format = "0.0%"


def build_master_data(ws):
    title(ws, "Master Data")
    set_layout(ws)
    section(ws, 3, "Household Information")
    ws["A4"] = "Primary Name"; ws["B4"] = "Your Name"; input_cell(ws["B4"])
    ws["A5"] = "Spouse Name"; ws["B5"] = "Spouse Name"; input_cell(ws["B5"])
    ws["A6"] = "State"; ws["B6"] = "TX"; input_cell(ws["B6"])
    ws["A7"] = "Current Date"; ws["B7"] = "=TODAY()"; formula_cell(ws["B7"])

    section(ws, 9, "Children Information")
    headers = ["Child Name", "Birthdate", "Monthly Support", "Age Support Ends", "Expiration Date"]
    for i, h in enumerate(headers, 1):
        ws.cell(10, i, h).fill = PatternFill("solid", fgColor=COLORS["gray_fill"])
    for r in range(11, 14):
        ws.cell(r, 1, f"Child {r-10}"); input_cell(ws.cell(r, 1))
        ws.cell(r, 2, datetime(2012 + (r-11) * 2, 1, 1)); input_cell(ws.cell(r, 2)); ws.cell(r, 2).number_format = "mm/dd/yyyy"
        ws.cell(r, 3, 500); input_cell(ws.cell(r, 3)); currency(ws.cell(r, 3))
        ws.cell(r, 4, 18); input_cell(ws.cell(r, 4))
        ws.cell(r, 5, f"=EDATE(B{r},D{r}*12)"); formula_cell(ws.cell(r, 5)); ws.cell(r, 5).number_format = "mm/dd/yyyy"

    section(ws, 15, "Debt Names & Categories")
    ws["A16"] = "Debt Name"; ws["B16"] = "Category"
    for c in ["A16", "B16"]:
        ws[c].fill = PatternFill("solid", fgColor=COLORS["gray_fill"])
    debt_rows = [("Credit Card 1", "Credit Card"), ("Credit Card 2", "Credit Card"), ("Boat Loan", "Boat"), ("Car Loan", "Auto"), ("Personal Loan", "Personal")]
    for i, (n, cat) in enumerate(debt_rows, 17):
        ws[f"A{i}"] = n; input_cell(ws[f"A{i}"])
        ws[f"B{i}"] = cat; input_cell(ws[f"B{i}"])

    section(ws, 23, "Lookup Lists")
    pay = ["Weekly", "Biweekly", "Semimonthly", "Monthly"]
    for i, v in enumerate(pay, 24):
        ws[f"A{i}"] = v; input_cell(ws[f"A{i}"])
    scenarios = ["Conservative", "Moderate", "Aggressive"]
    for i, v in enumerate(scenarios, 24):
        ws[f"C{i}"] = v; input_cell(ws[f"C{i}"])


def build_dashboard(ws):
    title(ws, "Dashboard")
    set_layout(ws)
    cards = [
        ("Total Monthly Income", "='Income'!B34"),
        ("Total Monthly Expenses", "='Budget'!D80"),
        ("Monthly Surplus", "=B4-B5"),
        ("Total Debt Balance", "='Debt Tracker'!C30"),
        ("Net Worth", "='Net Worth Statement'!B31"),
        ("Emergency Fund Coverage", "='Emergency Fund Analysis'!B6"),
        ("Months Until Wife Can Retire", "='Retirement Planner'!B11"),
    ]
    r = 3
    for label, formula in cards:
        ws[f"A{r}"] = label
        ws[f"B{r}"] = formula
        formula_cell(ws[f"B{r}"], cross=True)
        currency(ws[f"B{r}"])
        r += 1
    ws["B6"].number_format = "0.0"
    ws["B9"].number_format = "0"
    section(ws, 12, "Critical Milestones")
    ws["A13"] = "Child 1 support expiration"; ws["B13"] = "='Master Data'!E11"
    ws["A14"] = "Child 2 support expiration"; ws["B14"] = "='Master Data'!E12"
    ws["A15"] = "Debt free date"; ws["B15"] = "='Debt Payoff Strategy'!B21"
    ws["A16"] = "Emergency fund funded date"; ws["B16"] = "='Emergency Fund Analysis'!B14"
    ws["A17"] = "Projected retirement date"; ws["B17"] = "='Retirement Planner'!B12"
    for r in range(13, 18):
        ws[f"B{r}"].number_format = "mm/dd/yyyy"
        formula_cell(ws[f"B{r}"], cross=True)


def build_paycheck(ws):
    title(ws, "Paycheck Calculator")
    set_layout(ws)
    section(ws, 3, "Inputs")
    labels = ["Gross Pay", "Pay Frequency", "Federal Tax", "State Tax", "Social Security", "Medicare", "Health Insurance", "401k Contribution", "Other 1", "Other 2", "Other 3"]
    for i, lbl in enumerate(labels, 4):
        ws[f"A{i}"] = lbl
        ws[f"B{i}"] = 0 if i != 5 else "Biweekly"
        input_cell(ws[f"B{i}"])
    ws["B4"] = 4500; currency(ws["B4"])
    mult_formula = '=IF(B5="Weekly",52,IF(B5="Biweekly",26,IF(B5="Semimonthly",24,12)))'
    ws["D4"] = "Pay Multiplier"; ws["E4"] = mult_formula; formula_cell(ws["E4"])

    section(ws, 16, "Calculations")
    calc = {
        17: ("Total Deductions", "=SUM(B6:B14)"),
        18: ("Net Pay", "=B4-B17"),
        19: ("Annual Gross", "=B4*$E$4"),
        20: ("Annual Federal", "=B6*$E$4"),
        21: ("Annual State", "=B7*$E$4"),
        22: ("Annual Social Security", "=B8*$E$4"),
        23: ("Annual Medicare", "=B9*$E$4"),
        24: ("Annual Health", "=B10*$E$4"),
        25: ("Annual 401k", "=B11*$E$4"),
        26: ("Annual Other Deductions", "=SUM(B12:B14)*$E$4"),
        27: ("Annual Total Deductions", "=SUM(B20:B26)"),
        28: ("Annual Net Pay", "=B19-B27"),
        29: ("Monthly Net Pay", "=IFERROR(B28/12,0)"),
        30: ("Effective Tax Rate", "=IFERROR((B20+B21+B22+B23)/B19,0)"),
    }
    for r, (l, f) in calc.items():
        ws[f"A{r}"] = l
        ws[f"B{r}"] = f
        formula_cell(ws[f"B{r}"])
        if r != 30:
            currency(ws[f"B{r}"])
    percent(ws["B30"])


def build_simple_cross(ws, title_text, rows):
    title(ws, title_text)
    set_layout(ws)
    r = 3
    for lbl, formula, kind in rows:
        ws[f"A{r}"] = lbl
        ws[f"B{r}"] = formula
        formula_cell(ws[f"B{r}"], cross="'" in formula)
        if kind == "currency":
            currency(ws[f"B{r}"])
        elif kind == "percent":
            percent(ws[f"B{r}"])
        elif kind == "date":
            ws[f"B{r}"].number_format = "mm/dd/yyyy"
        r += 1


def build_budget(ws):
    title(ws, "Budget")
    set_layout(ws)
    ws["A3"] = "Line Item"
    ws["B3"] = "Category"
    ws["C3"] = "Monthly Budget"
    ws["D3"] = "Month 1"
    ws["E3"] = "Month 2"
    ws["F3"] = "YTD"
    for c in ["A3", "B3", "C3", "D3", "E3", "F3"]:
        ws[c].fill = PatternFill("solid", fgColor=COLORS["gray_fill"])
    rows = [
        ("Mortgage or rent", "Housing"), ("Utilities electric", "Housing"), ("Car payment 1", "Transportation"),
        ("Groceries", "Food"), ("Cell phones", "Utilities"), ("Credit card 1 minimum", "Debt"),
        ("Health insurance", "Healthcare"), ("Emergency fund contribution", "Savings"), ("Entertainment", "Entertainment")
    ]
    r = 4
    for item, cat in rows:
        ws[f"A{r}"] = item
        ws[f"B{r}"] = cat
        ws[f"C{r}"] = 0; input_cell(ws[f"C{r}"]); currency(ws[f"C{r}"])
        ws[f"D{r}"] = 0; input_cell(ws[f"D{r}"]); currency(ws[f"D{r}"])
        ws[f"E{r}"] = 0; input_cell(ws[f"E{r}"]); currency(ws[f"E{r}"])
        ws[f"F{r}"] = f"=SUM(D{r}:E{r})"; formula_cell(ws[f"F{r}"])
        currency(ws[f"F{r}"])
        r += 1
    ws["A80"] = "Total Monthly Expenses"
    ws["D80"] = "=SUM(D4:D12)"
    formula_cell(ws["D80"])
    currency(ws["D80"])


def build_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    for s in SHEETS:
        wb.create_sheet(s)

    build_master_data(wb["Master Data"])
    build_dashboard(wb["Dashboard"])
    build_paycheck(wb["Paycheck Calculator"])

    build_simple_cross(wb["Tax Planning"], "Tax Planning", [
        ("Annual gross income", "='Paycheck Calculator'!B19", "currency"),
        ("Pre-tax deductions", "='Paycheck Calculator'!B25+'Paycheck Calculator'!B24", "currency"),
        ("AGI", "=B3-B4", "currency"),
        ("Standard deduction", "='Tax Assumptions'!B15", "currency"),
        ("Taxable income", "=MAX(0,B5-B6)", "currency"),
        ("Estimated federal tax", "=IFERROR(B7*'Tax Assumptions'!B6,0)", "currency"),
    ])

    build_simple_cross(wb["Income"], "Income", [
        ("Your monthly net salary", "='Paycheck Calculator'!B29", "currency"),
        ("VA Disability monthly", "0", "currency"),
        ("Spouse monthly net income", "0", "currency"),
        ("Child 1 name", "='Master Data'!A11", "text"),
        ("Child 1 monthly", "='Master Data'!C11", "currency"),
        ("Child 1 expiration", "='Master Data'!E11", "date"),
        ("Child 1 status", "=IF(TODAY()<B8,\"ACTIVE\",\"EXPIRED\")", "text"),
        ("Child 1 included", "=IF(B9=\"ACTIVE\",B7,0)", "currency"),
        ("Total monthly income", "=SUM(B3:B5,B10)", "currency"),
        ("Total annual income", "=B11*12", "currency"),
    ])
    wb["Income"]["B34"] = "=B11"; formula_cell(wb["Income"]["B34"]); currency(wb["Income"]["B34"])

    build_budget(wb["Budget"])

    build_simple_cross(wb["Debt Tracker"], "Debt Tracker", [
        ("Debt name", "='Master Data'!A17", "text"),
        ("Debt category", "=IFERROR(VLOOKUP(B3,'Master Data'!A17:B30,2,FALSE),\"Not Found\")", "text"),
        ("Current balance", "15000", "currency"),
        ("APR", "0.189", "percent"),
        ("Minimum payment", "350", "currency"),
        ("Months to payoff minimum", "=IFERROR(ROUNDUP(NPER(B6/12,-B7,B5),0),\"Paid Off\")", "text"),
        ("Total interest minimum", "=IFERROR((B8*B7)-B5,0)", "currency"),
        ("Payoff date minimum", "=IFERROR(EDATE(TODAY(),B8),\"Paid Off\")", "date"),
        ("Extra payment", "500", "currency"),
        ("Months with extra", "=IFERROR(ROUNDUP(NPER(B6/12,-(B7+B11),B5),0),\"Paid Off\")", "text"),
    ])
    wb["Debt Tracker"]["C30"] = "=B5"; formula_cell(wb["Debt Tracker"]["C30"]); currency(wb["Debt Tracker"]["C30"])

    for name in SHEETS[7:]:
        ws = wb[name]
        if ws["A1"].value is None:
            title(ws, name)
            set_layout(ws)
            ws["A3"] = "Primary metric"
            ws["B3"] = "=0"
            formula_cell(ws["B3"])

    # Fill key formulas for downstream references
    rp = wb["Retirement Planner"]
    rp["A3"] = "Current nest egg"; rp["B3"] = 100000; input_cell(rp["B3"]); currency(rp["B3"])
    rp["A4"] = "Monthly invested"; rp["B4"] = 2500; input_cell(rp["B4"]); currency(rp["B4"])
    rp["A5"] = "Annual return"; rp["B5"] = "='Settings'!B4"; formula_cell(rp["B5"], cross=True); percent(rp["B5"])
    rp["A6"] = "Target annual income"; rp["B6"] = 90000; input_cell(rp["B6"]); currency(rp["B6"])
    rp["A7"] = "Safe withdrawal rate"; rp["B7"] = "='Settings'!B5"; formula_cell(rp["B7"], cross=True); percent(rp["B7"])
    rp["A8"] = "Target nest egg"; rp["B8"] = "=IFERROR(B6/B7,0)"; formula_cell(rp["B8"]); currency(rp["B8"])
    rp["A11"] = "Months to retirement"; rp["B11"] = "=IFERROR(ROUNDUP(NPER(B5/12,B4,-B3,B8),0),\"Check Inputs\")"; formula_cell(rp["B11"])
    rp["A12"] = "Retirement date"; rp["B12"] = "=EDATE(TODAY(),B11)"; formula_cell(rp["B12"]); rp["B12"].number_format = "mm/dd/yyyy"

    ef = wb["Emergency Fund Analysis"]
    ef["A3"] = "Current balance"; ef["B3"] = 25000; input_cell(ef["B3"]); currency(ef["B3"])
    ef["A4"] = "Monthly expenses"; ef["B4"] = "='Budget'!D80"; formula_cell(ef["B4"], cross=True); currency(ef["B4"])
    ef["A6"] = "Months coverage"; ef["B6"] = "=IFERROR(B3/B4,0)"; formula_cell(ef["B6"])
    ef["A14"] = "Funded date"; ef["B14"] = "=EDATE(TODAY(),IFERROR(MAX(0,(B4*'Settings'!$B$6-B3)/1000),0))"; formula_cell(ef["B14"]); ef["B14"].number_format = "mm/dd/yyyy"

    nw = wb["Net Worth Statement"]
    nw["A30"] = "Total assets"; nw["B30"] = 350000; input_cell(nw["B30"]); currency(nw["B30"])
    nw["A31"] = "Net worth"; nw["B31"] = "=B30-120000"; formula_cell(nw["B31"]); currency(nw["B31"])

    st = wb["Settings"]
    st["A3"] = "Inflation rate"; st["B3"] = 0.025
    st["A4"] = "Expected market return"; st["B4"] = 0.07
    st["A5"] = "Safe withdrawal rate"; st["B5"] = 0.04
    st["A6"] = "Emergency fund target months"; st["B6"] = 6
    st["A7"] = "Extra debt payment"; st["B7"] = 500
    st["A8"] = "Years to project"; st["B8"] = 25
    for r in range(3, 9):
        input_cell(st[f"B{r}"])
        st[f"B{r}"].fill = PatternFill("solid", fgColor=COLORS["yellow_fill"])
    percent(st["B3"]); percent(st["B4"]); percent(st["B5"])

    tax = wb["Tax Assumptions"]
    tax["A3"] = "Current year"; tax["B3"] = "=YEAR(TODAY())"; formula_cell(tax["B3"])
    tax["A6"] = "Federal effective placeholder"; tax["B6"] = 0.12; input_cell(tax["B6"]); percent(tax["B6"])
    tax["A15"] = "Standard deduction MFJ"; tax["B15"] = 29200; input_cell(tax["B15"]); currency(tax["B15"])

    # Named ranges
    dn = [
        ("AnnualIncome", "'Income'!$B$12"),
        ("MonthlyExpenses", "'Budget'!$D$80"),
        ("TotalDebt", "'Debt Tracker'!$C$30"),
        ("CurrentNestEgg", "'Retirement Planner'!$B$3"),
        ("RetirementDate", "'Retirement Planner'!$B$12"),
        ("VADisabilityMonthly", "'Income'!$B$4"),
        ("EmergencyFundBalance", "'Emergency Fund Analysis'!$B$3"),
    ]
    for name, ref in dn:
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    # Data validation lists and names
    wb.defined_names.add(DefinedName("PayFreqList", attr_text="'Master Data'!$A$24:$A$27"))
    wb.defined_names.add(DefinedName("ScenarioList", attr_text="'Master Data'!$C$24:$C$26"))
    dv_pay = DataValidation(type="list", formula1="=PayFreqList")
    wb["Paycheck Calculator"].add_data_validation(dv_pay)
    dv_pay.add("B5")

    dv_sc = DataValidation(type="list", formula1="=ScenarioList")
    wb["Scenario Modeling"].add_data_validation(dv_sc)
    dv_sc.add("B4")

    dv_status = DataValidation(type="list", formula1='"Complete,Pending,Not Applicable"')
    wb["Monthly Update Checklist"].add_data_validation(dv_status)
    dv_status.add("D4:D20")

    # Conditional formats
    ef.conditional_formatting.add("B6", CellIsRule(operator="lessThan", formula=["3"], fill=PatternFill("solid", fgColor="00FF0000")))
    ef.conditional_formatting.add("B6", CellIsRule(operator="between", formula=["3", "6"], fill=PatternFill("solid", fgColor="00FFFF00")))
    ef.conditional_formatting.add("B6", CellIsRule(operator="greaterThan", formula=["6"], fill=PatternFill("solid", fgColor="0000FF00")))

    dt = wb["Debt Tracker"]
    dt.conditional_formatting.add("B6:B20", CellIsRule(operator="greaterThan", formula=["0.2"], fill=PatternFill("solid", fgColor="00FF9999")))

    inc = wb["Income"]
    inc.conditional_formatting.add("B9", FormulaRule(formula=['=AND(B8-TODAY()<365,TODAY()<B8)'], fill=PatternFill("solid", fgColor="00FFFF00")))

    # Tab colors
    for s in ["Income", "Paycheck Calculator", "Tax Planning", "Tax Assumptions"]:
        wb[s].sheet_properties.tabColor = TAB_COLORS["blue"]
    for s in ["Debt Tracker", "Debt Payoff Strategy", "Debt vs Investment Analysis"]:
        wb[s].sheet_properties.tabColor = TAB_COLORS["red"]
    for s in ["Emergency Fund Analysis", "200K Allocation Plan", "Tax-Optimized Savings Strategy"]:
        wb[s].sheet_properties.tabColor = TAB_COLORS["green"]
    for s in ["Retirement Planner", "Scenario Modeling", "Child Support Projection", "Child Support Impact Analysis"]:
        wb[s].sheet_properties.tabColor = TAB_COLORS["purple"]
    for s in ["Dashboard", "Net Worth Statement", "Master Data"]:
        wb[s].sheet_properties.tabColor = TAB_COLORS["gold"]
    for s in ["Settings", "Monthly Update Checklist"]:
        wb[s].sheet_properties.tabColor = TAB_COLORS["gray"]

    for ws in wb.worksheets:
        ws.print_options.gridLines = False
        ws.page_setup.fitToWidth = 1
        ws.oddHeader.center.text = f"&[Tab] - {OUTPUT}"
        ws.oddFooter.center.text = "Printed &[Date] | Page &[Page]"

    wb.save(OUTPUT)


if __name__ == "__main__":
    build_workbook()
    print(f"Generated {OUTPUT}")
