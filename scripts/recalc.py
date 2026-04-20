from __future__ import annotations

from openpyxl import load_workbook

FILE = "Personal Financial Master Plan.xlsx"


def main() -> int:
    wb = load_workbook(FILE, data_only=False)
    errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=#"):
                    errors.append(f"{ws.title}!{cell.coordinate} has hardcoded formula error token")
                if isinstance(value, str) and value in {"#DIV/0!", "#N/A", "#VALUE!", "#REF!"}:
                    errors.append(f"{ws.title}!{cell.coordinate} has error literal")

    if errors:
        print("status: failed")
        for err in errors:
            print(err)
        return 1

    print("status: success")
    print(f"checked_sheets: {len(wb.worksheets)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
